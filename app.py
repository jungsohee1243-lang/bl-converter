import io
import re
from datetime import datetime
from typing import List, Dict

import pandas as pd
import pdfplumber
import streamlit as st

PORT_MAP = {
    "SHIDAO CHINA": "CNSHD",
    "INCHEON,KOREA": "KRINC",
    "GUNSAN,KOREA": "KRKUV",
}

COLUMNS = [
    "비엘", "쉬퍼", "쉬퍼주소", "컨사이니", "컨사이니 사업자번호", "컨사이니 주소",
    "선명", "항차", "출발지", "도착지", "품명", "마크", "수량", "중량", "CBM", "원본파일명", "확인필요"
]


def group_lines(words, y_tol=3) -> List[str]:
    words = sorted(words, key=lambda w: (w["top"], w["x0"]))
    lines = []
    for w in words:
        txt = str(w.get("text", "")).strip()
        if not txt:
            continue
        for line in lines:
            if abs(line["top"] - w["top"]) <= y_tol:
                line["words"].append(w)
                line["top"] = sum(x["top"] for x in line["words"]) / len(line["words"])
                break
        else:
            lines.append({"top": w["top"], "words": [w]})

    result = []
    for line in sorted(lines, key=lambda l: l["top"]):
        ws = sorted(line["words"], key=lambda w: w["x0"])
        result.append(" ".join(w["text"] for w in ws).strip())
    return [x for x in result if x]


def words_in(words, x0, y0, x1, y1):
    return [
        w for w in words
        if w["x0"] >= x0 and w["x1"] <= x1 and w["top"] >= y0 and w["bottom"] <= y1
    ]


def comma_break(text: str) -> str:
    text = (text or "").strip()
    text = re.sub(r",\s*", ",\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_bl(full_text: str) -> str:
    candidates = re.findall(r"\b[A-Z]{2,}[A-Z0-9]{8,}\b", full_text)
    for c in candidates:
        if any(ch.isdigit() for ch in c) and len(c) >= 10 and "HAOKUNDA" not in c:
            return c
    return ""


def remove_bl(line: str, bl: str) -> str:
    if bl:
        line = re.sub(r"\b" + re.escape(bl) + r"\b", "", line)
    return line.strip()


def split_consignee_first_line(first_line: str):
    first_line = first_line.strip()
    biz = ""
    m_biz = re.search(r"\b\d{3}-\d{2}-\d{5}\b", first_line)
    if m_biz:
        biz = m_biz.group()
        name = first_line.replace(biz, "").strip()
        return name, biz, ""

    # 전화번호가 회사명 첫 줄에 붙은 케이스: ICHEON TRADE CO., LTD. 010-7681-8554
    m_phone = re.search(r"\b0\d{1,2}-?\d{3,4}-?\d{4}\b", first_line)
    if m_phone and "CO" in first_line.upper():
        name = first_line[:m_phone.start()].strip()
        rest = first_line[m_phone.start():].strip()
        return name, "", rest

    return first_line, "", ""


def parse_pdf(file_bytes: bytes, filename: str) -> Dict[str, str]:
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        page = pdf.pages[0]
        words = page.extract_words(x_tolerance=1, y_tolerance=3, keep_blank_chars=False)
        full_text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""

    bl = extract_bl(full_text)

    # 상단 좌측 박스 기준: 실제 BL 양식에서 쉬퍼/컨사이니는 항상 좌측 상단 구역에 있음.
    ship_lines = group_lines(words_in(words, 0, 15, 370, 85))
    ship_lines = [remove_bl(x, bl) for x in ship_lines]
    ship_lines = [x for x in ship_lines if x and x.upper() not in {"SHIPPER"}]
    shipper = ship_lines[0] if ship_lines else ""
    shipper_addr = "\n".join(ship_lines[1:])

    cons_lines = group_lines(words_in(words, 0, 85, 370, 180))
    cons_lines = [remove_bl(x, bl) for x in cons_lines]
    cons_lines = [x for x in cons_lines if x and x.upper() not in {"CONSIGNEE"}]
    consignee = ""
    consignee_biz = ""
    consignee_extra = ""
    if cons_lines:
        consignee, consignee_biz, consignee_extra = split_consignee_first_line(cons_lines[0])
    consignee_detail_lines = []
    if consignee_extra:
        consignee_detail_lines.append(consignee_extra)
    consignee_detail_lines += cons_lines[1:]
    consignee_addr = "\n".join(consignee_detail_lines)

    # 선명/항차/출발지
    vessel = voyage = pol = pod = ""
    text_lines = [x.strip() for x in full_text.split("\n") if x.strip()]
    for line in text_lines:
        if "SHIDAO CHINA" in line and ("PEARL" in line or re.search(r"\d{4,}[A-Z]", line)):
            before = line.split("SHIDAO CHINA")[0].strip()
            parts = before.split()
            if len(parts) >= 2:
                voyage = parts[-1]
                vessel = " ".join(parts[:-1])
            pol = "SHIDAO CHINA"
            break
    for line in text_lines:
        if re.fullmatch(r"(INCHEON|GUNSAN),KOREA", line):
            pod = line
            break

    # 마크: 좌측 하단 Marks & Nos. 박스에서 먼저 추출. 품명에서 밀리는 문제 방지.
    mark_lines_raw = group_lines(words_in(words, 0, 330, 170, 620))
    marks = []
    for line in mark_lines_raw:
        if re.search(r"Marks|Container|Seal|No\.", line, re.I):
            continue
        if re.search(r"[A-Z]{1,6}\d|\d+-\d+|~", line):
            marks.append(line.strip())
        elif line.strip() == "-" and marks:
            marks.append(line.strip())

    if not marks:
        marks = re.findall(r"\b[A-Z]{1,6}\d+[A-Z0-9]*[-~][A-Z0-9~\-]+\b", full_text)
    mark_set = set(marks)

    # 수량/중량/CBM
    qty = ""
    m_qty = re.search(r"\b(\d+)\s*(PKGS|PKG|CTNS|CTN|CARTONS|CARTON|BOXES|BOX)\b", full_text, re.I)
    if m_qty:
        qty = m_qty.group(1)

    weight = ""
    m_weight = re.search(r"\b([\d,]+(?:\.\d+)?)\s*KGS\b", full_text, re.I)
    if m_weight:
        weight = m_weight.group(1).replace(",", "")

    cbm = ""
    m_cbm = re.search(r"\b([\d,]+(?:\.\d+)?)\s*CBM\b", full_text, re.I)
    if m_cbm:
        cbm = m_cbm.group(1).replace(",", "")

    # 품명: Description 박스에서 SAID TO CONTAIN 아래만 추출. 좌측 마크는 제외.
    desc_lines_raw = group_lines(words_in(words, 170, 330, 470, 620))
    desc_lines = []
    started = False
    for line in desc_lines_raw:
        if "SAID TO CONTAIN" in line:
            started = True
            continue
        if "FREIGHT" in line or "SURRENDER" in line or "SHIPPED" in line:
            break
        if not started:
            continue
        if "SHIPPER'S LOAD" in line:
            continue
        if line.strip() in mark_set:
            continue
        if re.fullmatch(r"[-–—]", line.strip()):
            continue
        if re.fullmatch(r"[A-Z]{1,6}\d+[A-Z0-9]*[-~][A-Z0-9~\-]+", line.strip()):
            continue
        desc_lines.append(line.strip())

    # fallback: 텍스트 순서 기준. 단, 마크 패턴은 제외.
    if not desc_lines:
        started = False
        for line in text_lines:
            if "SAID TO CONTAIN" in line:
                started = True
                rem = line.split("SAID TO CONTAIN", 1)[1].strip()
                if rem:
                    desc_lines.append(rem)
                continue
            if started:
                if "FREIGHT" in line:
                    break
                if line in mark_set or re.fullmatch(r"[-–—]", line):
                    continue
                if re.fullmatch(r"[A-Z]{1,6}\d+[A-Z0-9]*[-~][A-Z0-9~\-]+", line):
                    continue
                desc_lines.append(line)

    result = {
        "비엘": bl,
        "쉬퍼": shipper,
        "쉬퍼주소": comma_break(shipper_addr),
        "컨사이니": consignee,
        "컨사이니 사업자번호": consignee_biz,
        "컨사이니 주소": comma_break(consignee_addr),
        "선명": vessel,
        "항차": voyage,
        "출발지": PORT_MAP.get(pol, pol),
        "도착지": PORT_MAP.get(pod, pod),
        "품명": "\n".join(desc_lines),
        "마크": "\n".join(marks),
        "수량": qty,
        "중량": weight,
        "CBM": cbm,
        "원본파일명": filename,
    }

    required = ["비엘", "쉬퍼", "컨사이니", "선명", "항차", "출발지", "도착지", "품명", "마크", "수량", "중량", "CBM"]
    missing = [k for k in required if not result.get(k)]
    result["확인필요"] = "" if not missing else "확인필요: " + ", ".join(missing)
    return result


def make_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="변환결과")
        ws = writer.book["변환결과"]
        ws.freeze_panes = "A2"
        widths = {
            "A": 18, "B": 28, "C": 42, "D": 24, "E": 18, "F": 44,
            "G": 18, "H": 10, "I": 12, "J": 12, "K": 42, "L": 24,
            "M": 10, "N": 12, "O": 10, "P": 28, "Q": 38,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        error_fill = PatternFill("solid", fgColor="FFF2CC")
        thin = Side(style="thin", color="D9D9D9")
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=len(COLUMNS)).value:
                for col in range(1, len(COLUMNS) + 1):
                    ws.cell(row=row, column=col).fill = error_fill
    return output.getvalue()


st.set_page_config(page_title="BL PDF 변환기", layout="wide")
st.title("BL PDF → 신고용 엑셀 변환기 v19")
st.caption("좌측 박스 기준 추출 + 마크/품명 분리 보강 버전")

uploaded_files = st.file_uploader("PDF 파일을 여러 개 업로드하세요", type=["pdf"], accept_multiple_files=True)

if uploaded_files:
    rows = []
    for file in uploaded_files:
        try:
            rows.append(parse_pdf(file.getvalue(), file.name))
        except Exception as e:
            rows.append({c: "" for c in COLUMNS} | {"원본파일명": file.name, "확인필요": f"변환실패: {e}"})

    df = pd.DataFrame(rows)
    for c in COLUMNS:
        if c not in df.columns:
            df[c] = ""
    df = df[COLUMNS]

    st.subheader("변환 결과 미리보기")
    st.dataframe(df, use_container_width=True)

    excel_bytes = make_excel(df)
    filename = "BL_PDF_변환결과_" + datetime.now().strftime("%Y%m%d_%H%M%S") + ".xlsx"
    st.download_button(
        "엑셀 다운로드",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
else:
    st.info("PDF 파일을 업로드하면 자동으로 변환 결과가 표시됩니다.")
